import sys
import traceback
import os

def _setup_crash_logger():
    """Write uncaught exceptions to a file on Android for debugging."""
    try:
        log_paths = [
            "/sdcard/cone_crash.log",
            "/storage/emulated/0/cone_crash.log",
            os.path.join(os.getcwd(), "cone_crash.log"),
        ]
        def handle_exception(exc_type, exc_value, exc_tb):
            msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_tb))
            for path in log_paths:
                try:
                    with open(path, 'w') as f:
                        f.write(msg)
                    break
                except Exception:
                    continue
            sys.__excepthook__(exc_type, exc_value, exc_tb)
        sys.excepthook = handle_exception
    except Exception:
        pass

_setup_crash_logger()

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.spinner import Spinner
from kivy.uix.widget import Widget
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.graphics import (Color, RoundedRectangle, Rectangle, Ellipse, Line)
from kivy.metrics import dp
from kivy.core.window import Window
from kivy.animation import Animation
from datetime import datetime
from kivy.clock import Clock
import os
import json

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── Palette ───────────────────────────────────────────────────────────────────
# Window.clearcolor moved to build()

C_BG        = (0.11, 0.09, 0.20, 1)
C_SURFACE   = (0.18, 0.14, 0.30, 1)
C_GLASS     = (0.22, 0.17, 0.38, 0.85)
C_PRIMARY   = (0.55, 0.40, 0.95, 1)
C_ACCENT    = (0.75, 0.55, 1.00, 1)
C_SUCCESS   = (0.30, 0.85, 0.65, 1)
C_DANGER    = (0.95, 0.35, 0.50, 1)
C_TEXT      = (0.92, 0.90, 1.00, 1)
C_MUTED     = (0.55, 0.50, 0.72, 1)
C_RESET     = (0.90, 0.40, 0.60, 1)
C_EDIT      = (0.35, 0.65, 0.95, 1)
C_WARN      = (0.95, 0.75, 0.20, 1)
C_RESULT    = (0.20, 0.75, 0.55, 1)
C_GOLD      = (1.00, 0.82, 0.20, 1)
C_PANEL_BG  = (0.08, 0.06, 0.16, 1)   # darker bg for side panel


# ── Canvas helpers ────────────────────────────────────────────────────────────
def _draw_glass_card(instance, value, radius=dp(18), color=None):
    c = color or C_BG
    instance.canvas.before.clear()
    with instance.canvas.before:
        Color(*c)
        RoundedRectangle(pos=instance.pos, size=instance.size, radius=[radius])
        Color(0.55, 0.40, 0.95, 0.18)
        Line(
            rounded_rectangle=(
                instance.pos[0] + 1, instance.pos[1] + 1,
                instance.size[0] - 2, instance.size[1] - 2,
                radius
            ),
            width=1.2,
        )


def make_card(padding=dp(16), radius=dp(18), color=None):
    layout = BoxLayout(orientation='vertical', padding=padding, spacing=dp(8),
                       size_hint_y=None)
    layout.bind(minimum_height=layout.setter('height'))

    def _draw(inst, val):
        _draw_glass_card(inst, val, radius=radius, color=color)

    layout.bind(pos=_draw, size=_draw)
    return layout


def section_label(text, color=None, font_size=None, bold=False):
    lbl = Label(
        text=text,
        color=color or C_TEXT,
        font_size=font_size or dp(15),
        bold=bold,
        halign='left',
        valign='middle',
        size_hint_y=None,
        height=dp(32),
    )
    lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (val[0], None)))
    return lbl


def make_input(hint='', next_input=None):
    ti = TextInput(
        hint_text=hint,
        multiline=False,
        input_filter='int',
        font_size=dp(18),
        size_hint_y=None,
        height=dp(52),
        padding=[dp(14), dp(15)],
        background_color=(*C_SURFACE[:3], 1),
        foreground_color=C_TEXT,
        cursor_color=C_PRIMARY,
        hint_text_color=(*C_MUTED[:3], 0.55),
    )
    if next_input is not None:
        def on_text_validate(instance, _nxt=next_input):
            Clock.schedule_once(lambda dt: setattr(_nxt, 'focus', True), 0.05)
        ti.bind(on_text_validate=on_text_validate)
    return ti


def make_button(text, bg_color, text_color=None, height=dp(56)):
    tc = text_color or C_TEXT
    btn = Button(
        text=text,
        font_size=dp(16),
        bold=True,
        size_hint_y=None,
        height=height,
        background_normal='',
        background_color=(0, 0, 0, 0),
        color=tc,
    )

    def _draw(inst, val):
        inst.canvas.before.clear()
        with inst.canvas.before:
            Color(*bg_color)
            RoundedRectangle(pos=inst.pos, size=inst.size, radius=[dp(28)])
            Color(1, 1, 1, 0.08)
            RoundedRectangle(
                pos=(inst.pos[0] + dp(4), inst.pos[1] + inst.size[1] * 0.55),
                size=(inst.size[0] - dp(8), inst.size[1] * 0.38),
                radius=[dp(28)],
            )

    btn.bind(pos=_draw, size=_draw)
    return btn


def result_label(text='', color=None):
    return Label(
        text=text,
        color=color or C_TEXT,
        font_size=dp(19),
        bold=True,
        halign='center',
        size_hint_y=None,
        height=dp(38),
    )


# ── Core calculation ──────────────────────────────────────────────────────────
def calculate(length, width, height, extra):
    try:
        length, width, height, extra = (
            int(length), int(width), int(height), int(extra)
        )
    except Exception:
        return None
    if height < 0:
        return None
    total = sum((length - i) * (width - i) for i in range(height))
    return total + extra


# ── Excel helpers ─────────────────────────────────────────────────────────────
def _get_excel_path(folder, date_str=None):
    if date_str is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
    return os.path.join(folder, f"{date_str}.xlsx")


def _get_all_excel_files(folder):
    if not os.path.exists(folder):
        return []
    files = [f[:-5] for f in os.listdir(folder)
             if f.endswith(".xlsx") and len(f) == 15]
    return sorted(files, reverse=True)


def _load_records(file_path):
    records = []
    if not os.path.exists(file_path):
        return records
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                records.append(dict(zip(headers, row)))
        wb.close()
    except Exception:
        pass
    return records


def _save_records(file_path, records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["#", "Title", "Date",
               "Small Length", "Small Width", "Small Height", "Small Extra", "Small Result",
               "Big Length",   "Big Width",   "Big Height",   "Big Extra",   "Big Result",
               "Total"]

    header_fill  = PatternFill("solid", start_color="6B50C8")
    header_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="9980D4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill, cell.font  = header_fill, header_font
        cell.alignment, cell.border = header_align, border

    alt_fill    = PatternFill("solid", start_color="1E1535")
    normal_fill = PatternFill("solid", start_color="150F28")
    data_font   = Font(name="Calibri", size=10, color="D4CAFF")
    center_align = Alignment(horizontal="center", vertical="center")

    for ri, rec in enumerate(records, 2):
        fill = alt_fill if ri % 2 == 0 else normal_fill
        values = [ri - 1, rec.get("Title", ""), rec.get("Date", ""),
                  rec.get("Small Length", ""), rec.get("Small Width", ""),
                  rec.get("Small Height", ""), rec.get("Small Extra", ""),
                  rec.get("Small Result", ""),
                  rec.get("Big Length", ""),   rec.get("Big Width", ""),
                  rec.get("Big Height", ""),   rec.get("Big Extra", ""),
                  rec.get("Big Result", ""),   rec.get("Total", "")]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill, cell.font = fill, data_font
            cell.alignment, cell.border = center_align, border

    col_widths = [4, 18, 20, 12, 12, 12, 12, 13, 10, 10, 10, 10, 11, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    wb.save(file_path)
    wb.close()


def _records_equal(r1, r2):
    keys = ["Title", "Small Length", "Small Width", "Small Height", "Small Extra",
            "Big Length",  "Big Width",  "Big Height",  "Big Extra"]
    return all(str(r1.get(k, "")) == str(r2.get(k, "")) for k in keys)


# ── DATE PICKER POPUP ─────────────────────────────────────────────────────────
class DatePickerPopup(Popup):
    def __init__(self, date_list, on_date_selected, title_text="Select Date", **kwargs):
        super().__init__(
            title=title_text,
            title_color=C_ACCENT,
            title_size=dp(17),
            separator_color=C_PRIMARY,
            background_color=(*C_BG[:3], 0.97),
            size_hint=(0.92, 0.80),
            **kwargs,
        )
        self.on_date_selected = on_date_selected

        root = BoxLayout(orientation='vertical', padding=dp(14), spacing=dp(10))

        root.add_widget(Label(
            text="Choose a saved date:",
            color=C_ACCENT, font_size=dp(15), bold=True,
            size_hint_y=None, height=dp(36), halign='center',
        ))

        scroll = ScrollView(size_hint=(1, 1))
        inner  = BoxLayout(orientation='vertical', spacing=dp(8), size_hint_y=None)
        inner.bind(minimum_height=inner.setter('height'))
        scroll.add_widget(inner)

        if not date_list:
            inner.add_widget(Label(
                text="No saved records found.",
                color=C_DANGER, font_size=dp(14),
                size_hint_y=None, height=dp(40),
            ))
        else:
            for ds in date_list:
                btn = make_button(ds, C_SURFACE, C_TEXT)
                btn.bind(on_press=lambda b, d=ds: self._pick(d))
                inner.add_widget(btn)

        root.add_widget(scroll)
        cancel_btn = make_button("Cancel", C_RESET)
        cancel_btn.bind(on_press=self.dismiss)
        root.add_widget(cancel_btn)
        self.content = root

    def _pick(self, date_str):
        self.dismiss()
        self.on_date_selected(date_str)


# ── RECORD LIST POPUP ─────────────────────────────────────────────────────────
class RecordListPopup(Popup):
    def __init__(self, date_str, records, indices,
                 on_record_selected, on_record_deleted=None, **kwargs):
        super().__init__(
            title=f"Records \u2014 {date_str}",
            title_color=C_ACCENT, title_size=dp(16),
            separator_color=C_PRIMARY,
            background_color=(*C_BG[:3], 0.97),
            size_hint=(0.95, 0.88),
            **kwargs,
        )
        self.on_record_selected = on_record_selected
        self.on_record_deleted  = on_record_deleted

        root = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(10))
        root.add_widget(Label(
            text="Tap a record to edit  \u2022  \U0001F5D1 to delete",
            color=C_MUTED, font_size=dp(13), bold=True,
            size_hint_y=None, height=dp(28), halign='center',
        ))

        scroll = ScrollView(size_hint=(1, 1))
        inner  = BoxLayout(orientation='vertical', spacing=dp(8), size_hint_y=None)
        inner.bind(minimum_height=inner.setter('height'))
        scroll.add_widget(inner)

        for pos, (idx, rec) in enumerate(zip(indices, records)):
            fl   = FloatLayout(size_hint_y=None, height=dp(88))
            card = BoxLayout(
                orientation='vertical', spacing=dp(2),
                size_hint=(1, None), height=dp(82),
                padding=[dp(12), dp(8)],
                pos_hint={'x': 0, 'y': 0},
            )

            def _draw_card(inst, val, c=card):
                c.canvas.before.clear()
                with c.canvas.before:
                    Color(*C_SURFACE)
                    RoundedRectangle(pos=c.pos, size=c.size, radius=[dp(14)])
                    Color(*C_PRIMARY[:3], 0.25)
                    Line(
                        rounded_rectangle=(c.pos[0]+1, c.pos[1]+1,
                                           c.size[0]-2, c.size[1]-2, dp(14)),
                        width=1.1,
                    )
            card.bind(pos=_draw_card, size=_draw_card)

            card.add_widget(Label(
                text=f"[b]{rec.get('Title', '(no title)')}[/b]",
                markup=True, color=C_TEXT, font_size=dp(15),
                size_hint_y=None, height=dp(28),
                halign='left',
                text_size=(max(Window.width, 100) * 0.68, None),
            ))
            card.add_widget(Label(
                text=(f"Time: {str(rec.get('Date',''))[-8:]}   "
                      f"Total: {rec.get('Total','')}   "
                      f"S: {rec.get('Small Result','')}   "
                      f"B: {rec.get('Big Result','')}"),
                color=C_MUTED, font_size=dp(12),
                size_hint_y=None, height=dp(22),
                halign='left',
                text_size=(max(Window.width, 100) * 0.68, None),
            ))

            edit_overlay = Button(
                background_normal='', background_color=(0, 0, 0, 0),
                size_hint=(0.78, None), height=dp(82),
                pos_hint={'x': 0, 'y': 0},
            )
            edit_overlay.bind(
                on_press=lambda b, p=pos, i=idx, r=rec: self._pick(p, i, r)
            )

            del_btn = Button(
                text='\U0001F5D1', font_size=dp(20),
                size_hint=(None, None), width=dp(50), height=dp(50),
                pos_hint={'right': 1, 'top': 1},
                background_normal='', background_color=(0, 0, 0, 0),
                color=C_DANGER,
            )

            def _del_draw(inst, val, b=del_btn):
                b.canvas.before.clear()
                with b.canvas.before:
                    Color(*C_DANGER[:3], 0.2)
                    RoundedRectangle(pos=b.pos, size=b.size, radius=[dp(14)])
            del_btn.bind(pos=_del_draw, size=_del_draw)
            del_btn.bind(
                on_press=lambda b, p=pos, i=idx, r=rec: self._confirm_delete(p, i, r)
            )

            fl.add_widget(card)
            fl.add_widget(edit_overlay)
            fl.add_widget(del_btn)
            inner.add_widget(fl)

        root.add_widget(scroll)
        cancel_btn = make_button("Close", C_RESET)
        cancel_btn.bind(on_press=self.dismiss)
        root.add_widget(cancel_btn)
        self.content = root

    def _pick(self, pos, idx, rec):
        self.dismiss()
        self.on_record_selected(pos, idx, rec)

    def _confirm_delete(self, pos, idx, rec):
        title = rec.get('Title', '(no title)')
        cp = Popup(
            title="Delete Record?",
            title_color=C_DANGER, title_size=dp(16),
            separator_color=C_DANGER,
            background_color=(*C_BG[:3], 0.97),
            size_hint=(0.85, 0.32),
        )
        box = BoxLayout(orientation='vertical', padding=dp(14), spacing=dp(10))
        box.add_widget(Label(
            text=f"Delete  [b]{title}[/b]?",
            markup=True, color=C_TEXT, font_size=dp(15),
            halign='center', size_hint_y=None, height=dp(40),
        ))
        btn_row = GridLayout(cols=2, spacing=dp(10), size_hint_y=None, height=dp(54))
        yes_btn = make_button("Delete", C_DANGER)
        no_btn  = make_button("Cancel", C_MUTED)

        def do_delete(*a):
            cp.dismiss()
            self.dismiss()
            if self.on_record_deleted:
                self.on_record_deleted(idx)

        yes_btn.bind(on_press=do_delete)
        no_btn.bind(on_press=cp.dismiss)
        btn_row.add_widget(yes_btn)
        btn_row.add_widget(no_btn)
        box.add_widget(btn_row)
        cp.content = box
        cp.open()


# ── EDIT FIELDS POPUP ─────────────────────────────────────────────────────────
class EditFieldsPopup(Popup):
    def __init__(self, rec, on_confirm, **kwargs):
        super().__init__(
            title="Edit Record",
            title_color=C_ACCENT, title_size=dp(17),
            separator_color=C_PRIMARY,
            background_color=(*C_BG[:3], 0.97),
            size_hint=(0.95, 0.92),
            **kwargs,
        )
        self.rec        = rec
        self.on_confirm = on_confirm
        self.edit_inputs = {}

        root = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(10))

        root.add_widget(Label(
            text=(f"[b]{rec.get('Title','(no title)')}[/b]\n"
                  f"[color=9980D4]{rec.get('Date','')}[/color]"),
            markup=True, color=C_TEXT, font_size=dp(14),
            halign='center', size_hint_y=None, height=dp(48),
        ))

        root.add_widget(Label(
            text="Title", color=C_MUTED, font_size=dp(13),
            size_hint_y=None, height=dp(22), halign='left',
            text_size=(max(Window.width, 100) - dp(48), None),
        ))
        self.title_edit = TextInput(
            text=str(rec.get("Title", "")),
            multiline=False,
            font_size=dp(17),
            size_hint_y=None, height=dp(50),
            padding=[dp(12), dp(14)],
            background_color=(*C_SURFACE[:3], 1),
            foreground_color=C_TEXT,
            cursor_color=C_PRIMARY,
        )
        root.add_widget(self.title_edit)

        scroll = ScrollView(size_hint=(1, 1))
        inner  = BoxLayout(orientation='vertical', spacing=dp(10), size_hint_y=None)
        inner.bind(minimum_height=inner.setter('height'))
        scroll.add_widget(inner)

        section_inputs = {}
        for sec_name, prefix in [("Small Cone", "Small"), ("Big Cone", "Big")]:
            inner.add_widget(Label(
                text=sec_name, color=C_PRIMARY, font_size=dp(15), bold=True,
                size_hint_y=None, height=dp(28), halign='left',
                text_size=(max(Window.width, 100) - dp(48), None),
            ))
            section_inputs[sec_name] = []
            for field in ["Length", "Width", "Height", "Extra"]:
                key = f"{prefix} {field}"
                ti = TextInput(
                    text=str(rec.get(key, "") or ""),
                    multiline=False, input_filter='int',
                    font_size=dp(17), size_hint_y=None, height=dp(50),
                    padding=[dp(12), dp(14)],
                    background_color=(*C_SURFACE[:3], 1),
                    foreground_color=C_TEXT, cursor_color=C_PRIMARY,
                )
                self.edit_inputs[key] = ti
                section_inputs[sec_name].append((key, ti))

            grid = GridLayout(cols=2, spacing=[dp(10), dp(6)], size_hint_y=None)
            grid.bind(minimum_height=grid.setter('height'))
            for f, ti in section_inputs[sec_name]:
                short = f.split(" ")[1]
                cell  = BoxLayout(orientation='vertical', spacing=dp(2),
                                  size_hint_y=None, height=dp(74))
                cell.add_widget(Label(
                    text=short, color=C_MUTED, font_size=dp(13),
                    size_hint_y=None, height=dp(22), halign='left',
                    text_size=(max(Window.width, 100) * 0.40, None),
                ))
                cell.add_widget(ti)
                grid.add_widget(cell)
            inner.add_widget(grid)

        root.add_widget(scroll)

        btn_row = GridLayout(cols=2, spacing=dp(10), size_hint_y=None, height=dp(58))
        confirm_btn = make_button("Confirm Edit", C_SUCCESS)
        cancel_btn  = make_button("Cancel", C_RESET)
        confirm_btn.bind(on_press=self._confirm)
        cancel_btn.bind(on_press=self.dismiss)
        btn_row.add_widget(confirm_btn)
        btn_row.add_widget(cancel_btn)
        root.add_widget(btn_row)
        self.content = root

    def _confirm(self, *args):
        self.on_confirm(self.title_edit.text, self.edit_inputs)
        self.dismiss()


# ── RESULT DETAIL POPUP ───────────────────────────────────────────────────────
class ResultDetailPopup(Popup):
    """Clean detail: Title / Small / Big / Total — nothing else."""
    def __init__(self, rec, **kwargs):
        super().__init__(
            title=rec.get('Title', 'Result'),
            title_color=C_GOLD,
            title_size=dp(17),
            separator_color=C_GOLD,
            background_color=(*C_BG[:3], 0.98),
            size_hint=(0.88, 0.60),
            **kwargs,
        )
        root = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(14))

        title  = rec.get('Title', '')
        small  = rec.get('Small Result', '-')
        big    = rec.get('Big Result',   '-')
        total  = rec.get('Total',        '-')

        def _row(label, value, val_color):
            row = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(48),
                            padding=[dp(10), dp(4)])
            def _draw(inst, val, b=row):
                b.canvas.before.clear()
                with b.canvas.before:
                    Color(*C_SURFACE)
                    RoundedRectangle(pos=b.pos, size=b.size, radius=[dp(10)])
            row.bind(pos=_draw, size=_draw)
            row.add_widget(Label(
                text=label, color=C_MUTED, font_size=dp(14),
                halign='left', valign='middle', size_hint_x=0.45,
            ))
            row.add_widget(Label(
                text=str(value) if str(value).strip() else '-',
                color=val_color, font_size=dp(20), bold=True,
                halign='right', valign='middle', size_hint_x=0.55,
            ))
            return row

        root.add_widget(_row("Small", small, C_EDIT))
        root.add_widget(_row("Big",   big,   C_SUCCESS))
        root.add_widget(_row("Total", total, C_GOLD))

        close_btn = make_button("Close", C_RESET)
        close_btn.bind(on_press=self.dismiss)
        root.add_widget(close_btn)
        self.content = root


# ── RESULT RECORD LIST (view only) ────────────────────────────────────────────
class ResultRecordListPopup(Popup):
    """Clean list: Title  Small  Big  Total. Duplicates show x2 combined row."""

    def __init__(self, date_str, records, **kwargs):
        from collections import defaultdict
        count = len(records)
        super().__init__(
            title=f"{date_str}  ({count} records)",
            title_color=C_GOLD,
            title_size=dp(15),
            separator_color=C_GOLD,
            background_color=(*C_BG[:3], 0.98),
            size_hint=(0.95, 0.93),
            **kwargs,
        )

        # Count duplicates
        name_count = defaultdict(int)
        for rec in records:
            name_count[rec.get('Title', '')] += 1
        seen_dup = set()

        def _sum_field(name, field):
            return sum(
                int(r.get(field, 0) or 0)
                for r in records
                if r.get('Title', '') == name
                and str(r.get(field, '')).lstrip('-').isdigit()
            )

        root = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))

        # Column header row
        col_hdr = GridLayout(
            cols=4, size_hint_y=None, height=dp(28),
            spacing=[dp(4), 0], padding=[dp(14), 0],
        )
        for lbl_txt, col in [("Title", C_MUTED), ("Small", C_EDIT),
                              ("Big", C_SUCCESS), ("Total", C_GOLD)]:
            col_hdr.add_widget(Label(
                text=lbl_txt, color=col, font_size=dp(12), bold=True,
                halign='center', valign='middle',
            ))
        root.add_widget(col_hdr)

        scroll = ScrollView(size_hint=(1, 1))
        inner  = BoxLayout(orientation='vertical', spacing=dp(6), size_hint_y=None)
        inner.bind(minimum_height=inner.setter('height'))
        scroll.add_widget(inner)

        for rec in records:
            name  = rec.get('Title', '')
            count = name_count[name]
            is_dup = count > 1

            if is_dup:
                # Show one combined x2/x3 row the first time, skip all individuals
                if name in seen_dup:
                    continue
                seen_dup.add(name)
                sum_small = _sum_field(name, 'Small Result')
                sum_big   = _sum_field(name, 'Big Result')
                sum_total = _sum_field(name, 'Total')

                dup_row = GridLayout(
                    cols=4, size_hint_y=None, height=dp(46),
                    spacing=[dp(4), 0], padding=[dp(10), dp(4)],
                )

                def _draw_dup(inst, val, b=dup_row):
                    b.canvas.before.clear()
                    with b.canvas.before:
                        Color(*C_GOLD[:3], 0.12)
                        RoundedRectangle(pos=b.pos, size=b.size, radius=[dp(10)])
                        Color(*C_GOLD[:3], 0.40)
                        Line(rounded_rectangle=(*b.pos, *b.size, dp(10)), width=1.2)
                dup_row.bind(pos=_draw_dup, size=_draw_dup)

                dup_row.add_widget(Label(
                    text=f"[b]{name}[/b] x{count}",
                    markup=True, color=C_GOLD, font_size=dp(13),
                    halign='center', valign='middle',
                ))
                for val, col in [(sum_small, C_EDIT), (sum_big, C_SUCCESS),
                                 (sum_total, C_GOLD)]:
                    dup_row.add_widget(Label(
                        text=f"[b]{val}[/b]", markup=True,
                        color=col, font_size=dp(14),
                        halign='center', valign='middle',
                    ))
                inner.add_widget(dup_row)

            else:
                # Unique entry — show individual row normally
                small = rec.get('Small Result', '')
                big   = rec.get('Big Result', '')
                total = rec.get('Total', '')

                row = GridLayout(
                    cols=4, size_hint_y=None, height=dp(40),
                    spacing=[dp(4), 0], padding=[dp(10), dp(2)],
                )

                def _draw_row(inst, val, b=row):
                    b.canvas.before.clear()
                    with b.canvas.before:
                        Color(*C_SURFACE)
                        RoundedRectangle(pos=b.pos, size=b.size, radius=[dp(8)])
                row.bind(pos=_draw_row, size=_draw_row)

                row.add_widget(Label(
                    text=name or '(no title)',
                    color=C_TEXT, font_size=dp(13),
                    halign='center', valign='middle',
                ))
                for val, col in [(small, C_EDIT), (big, C_SUCCESS), (total, C_GOLD)]:
                    row.add_widget(Label(
                        text=str(val) if str(val).strip() else '-',
                        color=col, font_size=dp(14), bold=True,
                        halign='center', valign='middle',
                    ))
                inner.add_widget(row)

        root.add_widget(scroll)
        back_btn = make_button("Back", C_RESET)
        back_btn.bind(on_press=self.dismiss)
        root.add_widget(back_btn)
        self.content = root


# ── RESULT DATE PICKER ────────────────────────────────────────────────────────
class ResultViewDatePopup(Popup):
    """Date list popup for viewing results."""
    def __init__(self, folder, **kwargs):
        super().__init__(
            title="Saved Results",
            title_color=C_GOLD,
            title_size=dp(17),
            separator_color=C_GOLD,
            background_color=(*C_BG[:3], 0.98),
            size_hint=(0.93, 0.85),
            **kwargs,
        )
        self.folder = folder
        date_list   = _get_all_excel_files(folder)

        root = BoxLayout(orientation='vertical', padding=dp(14), spacing=dp(10))

        root.add_widget(Label(
            text="Select a date to view results",
            color=C_ACCENT, font_size=dp(14), bold=True,
            size_hint_y=None, height=dp(32), halign='center',
        ))

        scroll = ScrollView(size_hint=(1, 1))
        inner  = BoxLayout(orientation='vertical', spacing=dp(8), size_hint_y=None)
        inner.bind(minimum_height=inner.setter('height'))
        scroll.add_widget(inner)

        if not date_list:
            inner.add_widget(Label(
                text="No saved records found.\nSave some calculations first!",
                color=C_MUTED, font_size=dp(14),
                size_hint_y=None, height=dp(60), halign='center',
            ))
        else:
            for ds in date_list:
                btn = make_button(ds, C_SURFACE, C_TEXT, height=dp(52))
                btn.bind(on_press=lambda b, d=ds: self._open_date(d))
                inner.add_widget(btn)

        root.add_widget(scroll)
        cancel_btn = make_button("Cancel", C_RESET)
        cancel_btn.bind(on_press=self.dismiss)
        root.add_widget(cancel_btn)
        self.content = root

    def _open_date(self, date_str):
        self.dismiss()
        file_path = _get_excel_path(self.folder, date_str)
        records   = _load_records(file_path)
        if not records:
            return
        ResultRecordListPopup(date_str=date_str, records=records).open()


# ── SIDE PANEL (Drawer) ───────────────────────────────────────────────────────
class SidePanel(FloatLayout):
    """Slide-in panel from the left with menu items + settings."""

    PANEL_W = 0.78   # fraction of window width

    def __init__(self, calc_ref, **kwargs):
        super().__init__(**kwargs)
        self.calc_ref  = calc_ref
        self._open     = False
        self.size_hint = (1, 1)
        # FloatLayout itself is NEVER disabled — that would block all child touches.
        # Instead we disable the scrim + panel individually when closed.

        # ── Dark scrim overlay ────────────────────────────────────────────────
        self.scrim = Button(
            background_normal='', background_color=(0, 0, 0, 0),
            size_hint=(1, 1), pos_hint={'x': 0, 'y': 0},
            opacity=0,
            disabled=True,   # starts disabled; enabled on open()
        )
        self.scrim.bind(on_press=lambda *a: self.close())
        self.add_widget(self.scrim)

        # ── Panel body ────────────────────────────────────────────────────────
        panel_w = max(Window.width, 100) * self.PANEL_W
        self.panel = BoxLayout(
            orientation='vertical',
            size_hint=(None, 1),
            width=panel_w,
            x=-panel_w,          # starts off-screen
            disabled=True,       # starts disabled; enabled on open()
        )

        def _draw_panel(inst, val):
            inst.canvas.before.clear()
            with inst.canvas.before:
                Color(*C_PANEL_BG)
                Rectangle(pos=inst.pos, size=inst.size)
                Color(*C_PRIMARY[:3], 0.25)
                Line(points=[inst.pos[0] + inst.size[0],
                              inst.pos[1],
                              inst.pos[0] + inst.size[0],
                              inst.pos[1] + inst.size[1]], width=1.5)
        self.panel.bind(pos=_draw_panel, size=_draw_panel)

        # ── Panel header ──────────────────────────────────────────────────────
        ph = BoxLayout(
            orientation='vertical',
            size_hint_y=None, height=dp(110),
            padding=[dp(20), dp(28), dp(20), dp(10)],
        )
        def _draw_ph(inst, val):
            inst.canvas.before.clear()
            with inst.canvas.before:
                Color(*C_PRIMARY[:3], 0.15)
                Rectangle(pos=inst.pos, size=inst.size)
        ph.bind(pos=_draw_ph, size=_draw_ph)

        ph.add_widget(Label(
            text="Cone Calculator",
            color=C_ACCENT, font_size=dp(20), bold=True,
            halign='left', text_size=(panel_w - dp(40), None),
            size_hint_y=None, height=dp(34),
        ))
        ph.add_widget(Label(
            text="by Ken",
            color=C_PRIMARY, font_size=dp(13),
            halign='left', text_size=(panel_w - dp(40), None),
            size_hint_y=None, height=dp(22),
        ))
        ph.add_widget(Label(
            text="v 1.2.0",
            color=(*C_MUTED[:3], 0.55), font_size=dp(11),
            halign='left', text_size=(panel_w - dp(40), None),
            size_hint_y=None, height=dp(18),
        ))
        self.panel.add_widget(ph)

        # ── Divider ───────────────────────────────────────────────────────────
        self.panel.add_widget(Widget(size_hint_y=None, height=dp(1)))

        # ── Scroll area for menu items ─────────────────────────────────────────
        scroll = ScrollView(size_hint=(1, 1))
        menu   = BoxLayout(orientation='vertical', spacing=dp(4),
                           padding=[dp(12), dp(14), dp(12), dp(14)],
                           size_hint_y=None)
        menu.bind(minimum_height=menu.setter('height'))
        scroll.add_widget(menu)
        self.panel.add_widget(scroll)

        def _menu_btn(label_text, color, callback):
            btn = Button(
                text=label_text,
                font_size=dp(17),
                bold=True,
                color=color,
                halign='left',
                valign='middle',
                background_normal='',
                background_down='',
                background_color=(0, 0, 0, 0),
                size_hint_y=None,
                height=dp(58),
                padding=[dp(20), 0],
            )
            btn.text_size = (panel_w - dp(40), dp(58))

            def _draw(inst, val, c=color):
                inst.canvas.before.clear()
                with inst.canvas.before:
                    Color(*c[:3], 0.10)
                    RoundedRectangle(pos=inst.pos, size=inst.size, radius=[dp(12)])
                    Color(*c[:3], 0.28)
                    Line(rounded_rectangle=(*inst.pos, *inst.size, dp(12)), width=1.0)
            btn.bind(pos=_draw, size=_draw)
            btn.bind(on_press=lambda *a: callback())
            return btn

        # ── Menu items ────────────────────────────────────────────────────────
        menu.add_widget(_menu_btn(
            "Results",
            C_RESULT,
            lambda: (self.close(), self.calc_ref._open_results(None))
        ))

        menu.add_widget(Widget(size_hint_y=None, height=dp(4)))

        menu.add_widget(_menu_btn(
            "Edit Records",
            C_EDIT,
            lambda: (self.close(), self.calc_ref._edit(None))
        ))

        menu.add_widget(Widget(size_hint_y=None, height=dp(4)))

        menu.add_widget(_menu_btn(
            "Reset",
            C_RESET,
            lambda: (self.close(), self.calc_ref._reset(None))
        ))

        menu.add_widget(Widget(size_hint_y=None, height=dp(4)))

        menu.add_widget(_menu_btn(
            "Add Preset",
            C_SUCCESS,
            lambda: (self.close(), self.calc_ref._add_preset())
        ))

        menu.add_widget(Widget(size_hint_y=None, height=dp(4)))

        menu.add_widget(_menu_btn(
            "Remove Preset",
            C_DANGER,
            lambda: (self.close(), self.calc_ref._remove_preset())
        ))

        # Spacer
        menu.add_widget(Widget(size_hint_y=None, height=dp(16)))

        # ── Divider line ──────────────────────────────────────────────────────
        div2 = Widget(size_hint_y=None, height=dp(1))
        def _draw_d2(inst, val):
            inst.canvas.before.clear()
            with inst.canvas.before:
                Color(*C_PRIMARY[:3], 0.20)
                Rectangle(pos=(inst.pos[0] + dp(16), inst.pos[1]),
                          size=(inst.size[0] - dp(32), dp(1)))
        div2.bind(pos=_draw_d2, size=_draw_d2)
        menu.add_widget(div2)
        menu.add_widget(Widget(size_hint_y=None, height=dp(10)))

        # ── About section ─────────────────────────────────────────────────────
        settings_lbl = Label(
            text="ABOUT",
            color=(*C_MUTED[:3], 0.65), font_size=dp(11), bold=True,
            halign='left', text_size=(panel_w - dp(32), None),
            size_hint_y=None, height=dp(24),
        )
        menu.add_widget(settings_lbl)
        menu.add_widget(Widget(size_hint_y=None, height=dp(6)))

        info_items = [
            ("Version", "1.2.0"),
            ("Author", "Ken"),
            ("Save Format", "Excel (.xlsx)"),
            ("Storage", "/sdcard/MyResults"),
        ]
        for lbl_txt, val_txt in info_items:
            row = BoxLayout(
                orientation='horizontal', size_hint_y=None, height=dp(36),
                padding=[dp(14), dp(0)],
            )
            def _draw_row(inst, val):
                inst.canvas.before.clear()
                with inst.canvas.before:
                    Color(*C_SURFACE[:3], 0.40)
                    RoundedRectangle(pos=inst.pos, size=inst.size, radius=[dp(8)])
            row.bind(pos=_draw_row, size=_draw_row)
            row.add_widget(Label(
                text=lbl_txt, color=C_MUTED, font_size=dp(12),
                halign='left', text_size=(dp(120), None), size_hint_x=0.45,
            ))
            row.add_widget(Label(
                text=val_txt, color=C_TEXT, font_size=dp(12), bold=True,
                halign='right', text_size=(panel_w * 0.45, None), size_hint_x=0.55,
            ))
            menu.add_widget(row)
            menu.add_widget(Widget(size_hint_y=None, height=dp(4)))

        # ── Footer ────────────────────────────────────────────────────────────
        self.panel.add_widget(Label(
            text="\u2746  Created by Ken  \u2746",
            color=(*C_PRIMARY[:3], 0.40),
            font_size=dp(11), halign='center',
            size_hint_y=None, height=dp(32),
        ))

        self.add_widget(self.panel)

    def open(self):
        if self._open:
            return
        self._open  = True
        self.disabled = False
        self.scrim.disabled = False
        self.panel.disabled = False
        Animation(x=0, duration=0.22, t='out_cubic').start(self.panel)
        Animation(background_color=(0, 0, 0, 0.50), opacity=1,
                  duration=0.22).start(self.scrim)

    def close(self):
        if not self._open:
            return
        self._open = False
        panel_w = self.panel.width

        def _after_close(*a):
            self.scrim.disabled = True
            self.panel.disabled = True

        anim = Animation(x=-panel_w, duration=0.20, t='in_cubic')
        anim.bind(on_complete=_after_close)
        anim.start(self.panel)
        Animation(background_color=(0, 0, 0, 0), opacity=0,
                  duration=0.20).start(self.scrim)

    def toggle(self):
        if self._open:
            self.close()
        else:
            self.open()

    def on_touch_down(self, touch):
        # When closed, pass all touches through — never consume them
        if not self._open:
            return False
        return super().on_touch_down(touch)

    def on_touch_move(self, touch):
        if not self._open:
            return False
        return super().on_touch_move(touch)

    def on_touch_up(self, touch):
        if not self._open:
            return False
        return super().on_touch_up(touch)


# ── HAMBURGER BUTTON ──────────────────────────────────────────────────────────
class HamburgerButton(Button):
    """Three-stripe button like Claude AI top-left."""
    def __init__(self, **kwargs):
        super().__init__(
            background_normal='', background_color=(0, 0, 0, 0),
            size_hint=(None, None), width=dp(48), height=dp(48),
            text='',
            **kwargs,
        )
        self.bind(pos=self._redraw, size=self._redraw)

    def _redraw(self, inst, val):
        self.canvas.before.clear()
        with self.canvas.before:
            # Button background pill
            Color(*C_SURFACE)
            RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(14)])
            Color(*C_PRIMARY[:3], 0.35)
            Line(rounded_rectangle=(*self.pos, *self.size, dp(14)), width=1.2)
            # Three stripes
            Color(*C_ACCENT)
            cx = self.pos[0]
            cy = self.pos[1]
            w  = self.size[0]
            h  = self.size[1]
            stripe_w = w * 0.52
            sx = cx + (w - stripe_w) / 2
            for g in [0.64, 0.50, 0.36]:
                Line(
                    points=[sx, cy + h * g, sx + stripe_w, cy + h * g],
                    width=dp(2.5),
                )


# ── MAIN CALCULATOR WIDGET ────────────────────────────────────────────────────
class ConeCalculator(FloatLayout):

    DEFAULT_PRESETS = [
        "Select preset",
        "Red Zebra", "Blue Zebra", "Rose Greedem",
        "Green Apple", "Green Star", "Plain Tip",
        "Violet Triangle", "Rose Solid",
    ]

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self._last_saved_record = None
        # Load presets from disk (persistent), fall back to defaults
        self.PRESETS = self._load_presets()

        # Full-screen background
        with self.canvas.before:
            Color(0.11, 0.09, 0.20, 1)
            self._bg_rect = Rectangle(pos=self.pos, size=self.size)
        self.bind(pos=self._update_bg, size=self._update_bg)

        # ── Main content (scroll) ────────────────────────────────────────────
        scroll  = ScrollView(size_hint=(1, 1), do_scroll_x=False)
        content = BoxLayout(
            orientation='vertical',
            padding=[dp(16), dp(16), dp(16), dp(24)],
            spacing=dp(14),
            size_hint_y=None,
        )
        content.bind(minimum_height=content.setter('height'))
        scroll.add_widget(content)
        self.add_widget(scroll)

        # ── Header row ────────────────────────────────────────────────────────
        header_row = FloatLayout(size_hint_y=None, height=dp(72))

        # Hamburger — top LEFT only (no right button)
        self._hamburger = HamburgerButton(pos_hint={'x': 0, 'top': 1})
        self._hamburger.bind(on_press=self._toggle_panel)
        header_row.add_widget(self._hamburger)

        # Title block — offset from hamburger
        title_block = BoxLayout(
            orientation='vertical',
            size_hint=(None, None),
            width=max(Window.width, 100) - dp(68),
            height=dp(72),
            padding=[0, dp(6), 0, 0],
            pos_hint={'right': 1, 'top': 1},
        )
        title_block.add_widget(Label(
            text="Cone Calculator",
            color=C_ACCENT, font_size=dp(24), bold=True,
            halign='center', valign='bottom',
            text_size=(max(Window.width, 100) - dp(68), None),
            size_hint_y=None, height=dp(38),
        ))
        title_block.add_widget(Label(
            text="by Ken",
            color=C_PRIMARY, font_size=dp(13),
            halign='center', valign='top',
            text_size=(max(Window.width, 100) - dp(68), None),
            size_hint_y=None, height=dp(22),
        ))
        header_row.add_widget(title_block)

        content.add_widget(header_row)

        # ── Preset card ──────────────────────────────────────────────────────
        preset_card = make_card(color=C_BG)
        preset_card.add_widget(section_label("Preset", color=C_MUTED, bold=True))

        # Themed preset picker button
        self.preset_spinner = Button(
            text=self.PRESETS[0],
            font_size=dp(16),
            size_hint_y=None,
            height=dp(52),
            background_normal='',
            background_down='',
            background_color=(0, 0, 0, 0),
            color=C_TEXT,
            halign='left',
            padding=[dp(16), 0],
        )

        def _spin_draw(inst, val):
            inst.canvas.before.clear()
            with inst.canvas.before:
                Color(*C_SURFACE)
                RoundedRectangle(pos=inst.pos, size=inst.size, radius=[dp(12)])
                Color(*C_PRIMARY[:3], 0.50)
                Line(rounded_rectangle=(*inst.pos, *inst.size, dp(12)), width=1.4)
        self.preset_spinner.bind(pos=_spin_draw, size=_spin_draw)

        def _open_preset_popup(instance):
            # Build popup
            pop = Popup(
                title='Select Preset',
                title_color=C_ACCENT,
                title_size=dp(16),
                separator_color=C_PRIMARY,
                background_color=(*C_BG[:3], 0.97),
                size_hint=(0.88, 0.75),
            )

            root = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
            scroll = ScrollView(size_hint=(1, 1))
            inner  = BoxLayout(orientation='vertical', spacing=dp(6), size_hint_y=None)
            inner.bind(minimum_height=inner.setter('height'))
            scroll.add_widget(inner)

            for preset in self.PRESETS:
                is_selected = (preset == self.preset_spinner.text)
                btn_color   = C_PRIMARY if is_selected else C_SURFACE
                txt_color   = (1, 1, 1, 1) if is_selected else C_TEXT

                pb = Button(
                    text=preset,
                    size_hint_y=None,
                    height=dp(52),
                    font_size=dp(16),
                    bold=is_selected,
                    background_normal='',
                    background_down='',
                    background_color=(0, 0, 0, 0),
                    color=txt_color,
                )

                def _pb_draw(inst, val, bc=btn_color):
                    inst.canvas.before.clear()
                    with inst.canvas.before:
                        Color(*bc)
                        RoundedRectangle(pos=inst.pos, size=inst.size, radius=[dp(12)])
                        Color(*C_PRIMARY[:3], 0.30)
                        Line(rounded_rectangle=(*inst.pos, *inst.size, dp(12)), width=1.1)
                pb.bind(pos=_pb_draw, size=_pb_draw)

                def _pick(b, p=preset, pp=pop):
                    self.preset_spinner.text = p
                    self._on_preset(None, p)
                    pp.dismiss()
                pb.bind(on_release=_pick)
                inner.add_widget(pb)

            root.add_widget(scroll)
            cancel = make_button('Cancel', C_RESET, height=dp(50))
            cancel.bind(on_press=pop.dismiss)
            root.add_widget(cancel)
            pop.content = root
            pop.open()

        self.preset_spinner.bind(on_release=_open_preset_popup)
        preset_card.add_widget(self.preset_spinner)
        content.add_widget(preset_card)

        # ── Title card ───────────────────────────────────────────────────────
        title_card = make_card(color=C_BG)
        title_card.add_widget(section_label("Title", color=C_MUTED, bold=True))
        self.title_input = TextInput(
            hint_text="Enter title\u2026",
            multiline=False,
            font_size=dp(18),
            size_hint_y=None, height=dp(52),
            padding=[dp(14), dp(15)],
            background_color=(*C_SURFACE[:3], 1),
            foreground_color=C_TEXT,
            cursor_color=C_PRIMARY,
            hint_text_color=(*C_MUTED[:3], 0.55),
        )
        title_card.add_widget(self.title_input)
        content.add_widget(title_card)

        # ── Small Cone card ──────────────────────────────────────────────────
        small_card = make_card(color=C_BG)
        small_card.add_widget(section_label("Small Cone", color=C_PRIMARY,
                                            font_size=dp(16), bold=True))
        self.small_inputs = self._cone_fields(small_card)
        content.add_widget(small_card)

        # ── Big Cone card ────────────────────────────────────────────────────
        big_card = make_card(color=C_BG)
        big_card.add_widget(section_label("Big Cone", color=C_PRIMARY,
                                          font_size=dp(16), bold=True))
        self.big_inputs = self._cone_fields(big_card)
        content.add_widget(big_card)

        # Wire Enter key chain
        all_inputs = self.small_inputs + self.big_inputs

        def make_focus(nxt):
            def _go(*a):
                Clock.schedule_once(lambda dt: setattr(nxt, 'focus', True), 0.05)
            return _go

        self.title_input.bind(on_text_validate=make_focus(all_inputs[0]))
        for i, inp in enumerate(all_inputs[:-1]):
            inp.bind(on_text_validate=make_focus(all_inputs[i + 1]))

        # ── Results card ─────────────────────────────────────────────────────
        result_card = make_card(padding=dp(20), color=C_BG)
        result_card.add_widget(section_label("Results", color=C_MUTED, bold=True))
        self.lbl_small = result_label(color=C_EDIT)
        self.lbl_big   = result_label(color=C_SUCCESS)
        self.lbl_total = result_label(color=C_ACCENT)
        result_card.add_widget(self.lbl_small)
        result_card.add_widget(self.lbl_big)
        result_card.add_widget(self.lbl_total)
        content.add_widget(result_card)

        # ── Action buttons: Save | Reset ──────────────────────────────────────
        btn_row = GridLayout(cols=2, spacing=dp(10), size_hint_y=None, height=dp(60))
        save_btn  = make_button("Save",  C_PRIMARY)
        reset_btn = make_button("Reset", C_RESET)
        save_btn.bind(on_press=self._save)
        reset_btn.bind(on_press=self._reset)
        btn_row.add_widget(save_btn)
        btn_row.add_widget(reset_btn)
        content.add_widget(btn_row)

        # ── Status label ─────────────────────────────────────────────────────
        self.status_lbl = Label(
            text='', color=C_MUTED, font_size=dp(13),
            size_hint_y=None, height=dp(48),
            halign='center',
            text_size=(max(Window.width, 100) - dp(32), None),
        )
        content.add_widget(self.status_lbl)

        # ── Creator footer ───────────────────────────────────────────────────
        content.add_widget(Label(
            text="\u2746  Created by Ken  \u2746",
            color=(*C_PRIMARY[:3], 0.55),
            font_size=dp(12), halign='center',
            size_hint_y=None, height=dp(28),
        ))
        content.add_widget(Widget(size_hint_y=None, height=dp(16)))

        for inp in [*self.small_inputs, *self.big_inputs]:
            inp.bind(text=self._update)

        # ── Side panel (on top of everything) ────────────────────────────────
        self._side_panel = SidePanel(calc_ref=self, size_hint=(1, 1))
        self.add_widget(self._side_panel)

    def _toggle_panel(self, *args):
        self._side_panel.toggle()

    def _update_bg(self, instance, value):
        self._bg_rect.pos  = instance.pos
        self._bg_rect.size = instance.size

    def _cone_fields(self, parent):
        labels = ['Length', 'Width', 'Height', 'Extra']
        hints  = ['Length', 'Width', 'Height (steps)', 'Extra (offset)']
        inputs = []
        grid   = GridLayout(cols=2, spacing=[dp(10), dp(6)], size_hint_y=None)
        grid.bind(minimum_height=grid.setter('height'))

        for lbl_text, hint in zip(labels, hints):
            cell = BoxLayout(orientation='vertical', spacing=dp(2),
                             size_hint_y=None, height=dp(76))
            cell.add_widget(Label(
                text=lbl_text, color=C_MUTED, font_size=dp(13),
                size_hint_y=None, height=dp(22), halign='left',
                text_size=(max(Window.width, 100) * 0.40, None),
            ))
            inp = TextInput(
                hint_text=hint, multiline=False, input_filter='int',
                font_size=dp(17), size_hint_y=None, height=dp(50),
                padding=[dp(14), dp(15)],
                background_color=(*C_SURFACE[:3], 1),
                foreground_color=C_TEXT,
                cursor_color=C_PRIMARY,
                hint_text_color=(*C_MUTED[:3], 0.50),
            )
            cell.add_widget(inp)
            grid.add_widget(cell)
            inputs.append(inp)

        parent.add_widget(grid)
        return inputs

    def _update(self, *args):
        sl, sw, sh, se = [i.text for i in self.small_inputs]
        bl, bw, bh, be = [i.text for i in self.big_inputs]

        small = calculate(sl, sw, sh, se)
        big   = calculate(bl, bw, bh, be)

        _s = str(small) if small is not None else '—'
        self.lbl_small.text = f"Small = {_s}"
        _b = str(big) if big is not None else '—'
        self.lbl_big.text   = f"Big   = {_b}"

        if small is not None and big is not None:
            self.lbl_total.text = f"Total = {small + big}"
        else:
            self.lbl_total.text = "Total = \u2014"

    def _on_preset(self, spinner, value):
        if value and value != self.PRESETS[0]:
            self.title_input.text = value

    def _get_save_folder(self):
        paths = [
            "/sdcard/MyResults",
            "/storage/emulated/0/MyResults",
            os.path.join(os.getcwd(), "MyResults"),
        ]
        for folder in paths:
            try:
                os.makedirs(folder, exist_ok=True)
                test = os.path.join(folder, ".write_test")
                with open(test, 'w') as f:
                    f.write("ok")
                os.remove(test)
                return folder
            except Exception:
                continue
        return None

    def _build_record(self, name, now):
        sl, sw, sh, se = [i.text for i in self.small_inputs]
        bl, bw, bh, be = [i.text for i in self.big_inputs]
        small = calculate(sl, sw, sh, se)
        big   = calculate(bl, bw, bh, be)
        total = (small + big) if (small is not None and big is not None) else ""
        return {
            "Title": name, "Date": now,
            "Small Length": sl, "Small Width": sw,
            "Small Height": sh, "Small Extra": se,
            "Small Result": small if small is not None else "",
            "Big Length": bl,   "Big Width": bw,
            "Big Height": bh,   "Big Extra": be,
            "Big Result": big   if big   is not None else "",
            "Total": total,
        }

    def _save(self, *args):
        if not EXCEL_AVAILABLE:
            self.status_lbl.text  = "openpyxl not installed. Run: pip install openpyxl"
            self.status_lbl.color = C_DANGER
            return

        name = self.title_input.text.strip()
        if not name:
            self.status_lbl.text  = "Please enter a title first."
            self.status_lbl.color = C_DANGER
            return

        now   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        today = now[:10]

        folder = self._get_save_folder()
        if folder is None:
            self.status_lbl.text  = "Save failed. Check storage permissions."
            self.status_lbl.color = C_DANGER
            return

        file_path = _get_excel_path(folder, today)
        records   = _load_records(file_path)
        new_rec   = self._build_record(name, now)

        if self._last_saved_record and _records_equal(new_rec, self._last_saved_record):
            self.status_lbl.text  = "\u26A0 Duplicate save found \u2014 no changes made since last save."
            self.status_lbl.color = C_WARN
            return

        if any(_records_equal(new_rec, r) for r in records
               if str(r.get("Title", "")) == name):
            self.status_lbl.text  = "\u26A0 Duplicate save found \u2014 identical record already exists."
            self.status_lbl.color = C_WARN
            return

        records.append(new_rec)
        self._last_saved_record = new_rec

        try:
            _save_records(file_path, records)
            self.status_lbl.text  = f"\u2713 Saved to: {today}.xlsx"
            self.status_lbl.color = C_SUCCESS
        except Exception as e:
            self.status_lbl.text  = f"Save error: {e}"
            self.status_lbl.color = C_DANGER

    def _edit(self, *args):
        if not EXCEL_AVAILABLE:
            self.status_lbl.text  = "openpyxl not installed. Run: pip install openpyxl"
            self.status_lbl.color = C_DANGER
            return

        folder = self._get_save_folder()
        if folder is None:
            self.status_lbl.text  = "Cannot access storage."
            self.status_lbl.color = C_DANGER
            return

        date_list = _get_all_excel_files(folder)

        def on_date_selected(date_str):
            file_path = _get_excel_path(folder, date_str)
            records   = _load_records(file_path)
            if not records:
                self.status_lbl.text  = f"No records in {date_str}.xlsx"
                self.status_lbl.color = C_DANGER
                return
            indices = list(range(len(records)))

            def on_record_selected(pos, idx, rec):
                def on_confirm(new_title, edit_inputs):
                    changed   = False
                    new_title = new_title.strip()
                    if new_title and new_title != str(rec.get("Title", "")):
                        rec["Title"] = new_title
                        changed = True

                    field_map = {
                        "Small Length": "Small Length", "Small Width":  "Small Width",
                        "Small Height": "Small Height", "Small Extra":  "Small Extra",
                        "Big Length":   "Big Length",   "Big Width":    "Big Width",
                        "Big Height":   "Big Height",   "Big Extra":    "Big Extra",
                    }
                    for field, inp in edit_inputs.items():
                        new_val = inp.text.strip()
                        old_val = str(rec.get(field_map[field], "") or "")
                        if new_val != old_val:
                            rec[field_map[field]] = new_val
                            changed = True

                    if not changed:
                        self.status_lbl.text  = "\u26A0 No changes detected \u2014 record not updated."
                        self.status_lbl.color = C_WARN
                        return

                    sl = str(rec.get("Small Length", ""))
                    sw = str(rec.get("Small Width",  ""))
                    sh = str(rec.get("Small Height", ""))
                    se = str(rec.get("Small Extra",  ""))
                    bl = str(rec.get("Big Length",   ""))
                    bw = str(rec.get("Big Width",    ""))
                    bh = str(rec.get("Big Height",   ""))
                    be = str(rec.get("Big Extra",    ""))

                    small = calculate(sl, sw, sh, se)
                    big   = calculate(bl, bw, bh, be)
                    rec["Small Result"] = small if small is not None else ""
                    rec["Big Result"]   = big   if big   is not None else ""
                    rec["Total"] = (small + big) if (
                        small is not None and big is not None
                    ) else ""

                    records[idx] = rec
                    try:
                        _save_records(file_path, records)
                        self.status_lbl.text  = f"\u2713 Record updated in {date_str}.xlsx"
                        self.status_lbl.color = C_SUCCESS
                    except Exception as e:
                        self.status_lbl.text  = f"Edit save error: {e}"
                        self.status_lbl.color = C_DANGER

                EditFieldsPopup(rec=rec, on_confirm=on_confirm).open()

            def on_record_deleted(idx):
                try:
                    del records[idx]
                    _save_records(file_path, records)
                    self.status_lbl.text  = f"\u2713 Record deleted from {date_str}.xlsx"
                    self.status_lbl.color = C_SUCCESS
                except Exception as e:
                    self.status_lbl.text  = f"Delete error: {e}"
                    self.status_lbl.color = C_DANGER

            RecordListPopup(
                date_str=date_str, records=records, indices=indices,
                on_record_selected=on_record_selected,
                on_record_deleted=on_record_deleted,
            ).open()

        DatePickerPopup(date_list=date_list, on_date_selected=on_date_selected).open()

    def _open_results(self, *args):
        """Open the Results viewer."""
        folder = self._get_save_folder()
        if folder is None:
            self.status_lbl.text  = "Cannot access storage."
            self.status_lbl.color = C_DANGER
            return
        ResultViewDatePopup(folder=folder).open()

    # ── Preset persistence ────────────────────────────────────────────────────
    def _presets_path(self):
        for folder in ["/sdcard/MyResults", "/storage/emulated/0/MyResults",
                       os.path.join(os.getcwd(), "MyResults")]:
            try:
                os.makedirs(folder, exist_ok=True)
                test = os.path.join(folder, ".write_test")
                with open(test, 'w') as f:
                    f.write("ok")
                os.remove(test)
                return os.path.join(folder, "presets.json")
            except Exception:
                continue
        return None

    def _load_presets(self):
        path = self._presets_path()
        if path and os.path.exists(path):
            try:
                with open(path, 'r') as f:
                    data = json.load(f)
                if isinstance(data, list) and data:
                    # Always keep placeholder at index 0
                    if data[0] != "Select preset":
                        data.insert(0, "Select preset")
                    return data
            except Exception:
                pass
        return list(self.DEFAULT_PRESETS)

    def _save_presets(self):
        path = self._presets_path()
        if path:
            try:
                with open(path, 'w') as f:
                    json.dump(self.PRESETS, f)
            except Exception:
                pass

    def _reset(self, *args):
        for inp in [*self.small_inputs, *self.big_inputs]:
            inp.text = ''
        self.title_input.text    = ''
        self.preset_spinner.text = self.PRESETS[0]
        self.lbl_small.text      = ''
        self.lbl_big.text        = ''
        self.lbl_total.text      = ''
        self.status_lbl.text     = ''
        self._last_saved_record  = None

    def _add_preset(self):
        pop = Popup(
            title='Add Preset',
            title_color=C_SUCCESS,
            title_size=dp(16),
            separator_color=C_SUCCESS,
            background_color=(*C_BG[:3], 0.97),
            size_hint=(0.88, 0.32),
        )
        root = BoxLayout(orientation='vertical', padding=dp(16), spacing=dp(12))
        ti = TextInput(
            hint_text='Preset name...',
            multiline=False,
            font_size=dp(17),
            size_hint_y=None, height=dp(52),
            padding=[dp(14), dp(15)],
            background_color=(*C_SURFACE[:3], 1),
            foreground_color=C_TEXT,
            cursor_color=C_PRIMARY,
            hint_text_color=(*C_MUTED[:3], 0.55),
        )
        root.add_widget(ti)
        btn_row = GridLayout(cols=2, spacing=dp(10), size_hint_y=None, height=dp(52))
        add_btn    = make_button('Add', C_SUCCESS, height=dp(52))
        cancel_btn = make_button('Cancel', C_RESET, height=dp(52))

        def _do_add(*a):
            name = ti.text.strip()
            if name and name not in self.PRESETS:
                self.PRESETS.append(name)
                self._save_presets()
            pop.dismiss()

        add_btn.bind(on_press=_do_add)
        cancel_btn.bind(on_press=pop.dismiss)
        btn_row.add_widget(add_btn)
        btn_row.add_widget(cancel_btn)
        root.add_widget(btn_row)
        pop.content = root
        pop.open()

    def _remove_preset(self):
        # Only show removable presets (skip index 0 placeholder)
        removable = self.PRESETS[1:]
        if not removable:
            return

        pop = Popup(
            title='Remove Preset',
            title_color=C_DANGER,
            title_size=dp(16),
            separator_color=C_DANGER,
            background_color=(*C_BG[:3], 0.97),
            size_hint=(0.88, 0.75),
        )
        root   = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
        scroll = ScrollView(size_hint=(1, 1))
        inner  = BoxLayout(orientation='vertical', spacing=dp(6), size_hint_y=None)
        inner.bind(minimum_height=inner.setter('height'))
        scroll.add_widget(inner)

        for preset in removable:
            pb = Button(
                text=preset,
                size_hint_y=None, height=dp(52),
                font_size=dp(16),
                background_normal='', background_down='',
                background_color=(0, 0, 0, 0),
                color=C_TEXT,
            )

            def _pb_draw(inst, val):
                inst.canvas.before.clear()
                with inst.canvas.before:
                    Color(*C_SURFACE)
                    RoundedRectangle(pos=inst.pos, size=inst.size, radius=[dp(12)])
                    Color(*C_DANGER[:3], 0.35)
                    Line(rounded_rectangle=(*inst.pos, *inst.size, dp(12)), width=1.1)
            pb.bind(pos=_pb_draw, size=_pb_draw)

            def _pick(b, p=preset, pp=pop):
                if p in self.PRESETS:
                    self.PRESETS.remove(p)
                    self._save_presets()
                if self.preset_spinner.text == p:
                    self.preset_spinner.text = self.PRESETS[0]
                pp.dismiss()
            pb.bind(on_release=_pick)
            inner.add_widget(pb)

        root.add_widget(scroll)
        cancel_btn = make_button('Cancel', C_RESET, height=dp(50))
        cancel_btn.bind(on_press=pop.dismiss)
        root.add_widget(cancel_btn)
        pop.content = root
        pop.open()


# ── App entry point ───────────────────────────────────────────────────────────
class ConeApp(App):
    def build(self):
        self.title = 'Cone Calculator — by Ken'
        Window.clearcolor = (0.13, 0.10, 0.22, 1)
        self._request_android_permissions()
        return ConeCalculator()

    def _request_android_permissions(self):
        try:
            from android.permissions import request_permissions, Permission
            from android import api_version
            if api_version >= 23:
                request_permissions([
                    Permission.READ_EXTERNAL_STORAGE,
                    Permission.WRITE_EXTERNAL_STORAGE,
                ])
        except Exception:
            pass  # Not on Android or permission API unavailable


if __name__ == '__main__':
    ConeApp().run()
